from PyQt5.QtWidgets import QDialog, QTableWidget, QTableWidgetItem, QVBoxLayout, QHeaderView, QPushButton, QFileDialog, QMessageBox
import openpyxl
from datetime import datetime

class HistoryModal(QDialog):
    def __init__(self, parent=None, history_list=[]):
        super().__init__(parent)
        self.history_list = history_list
        self.setWindowTitle("Historial de la sesión")
        self.setGeometry(200, 200, 900, 700)

        self.layout = QVBoxLayout()

        self.table_widget = QTableWidget(self)
        self.table_widget.setColumnCount(5)
        self.table_widget.setHorizontalHeaderLabels(["Número de Trabajo", "Año", "Estado Anterior", "Estado Nuevo", "Fecha y Hora"])
        self.table_widget.setSizeAdjustPolicy(QTableWidget.AdjustToContents)
        self.table_widget.horizontalHeader().setStretchLastSection(True)
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        self.save_button = QPushButton("Guardar como Excel")
        self.save_button.setEnabled(False)
        self.save_button.clicked.connect(self.save_to_excel)

        self.populate_table()
        self.update_save_button_state()

        self.layout.addWidget(self.table_widget)
        self.layout.addWidget(self.save_button)
        self.setLayout(self.layout)

    def populate_table(self):
        self.table_widget.setRowCount(len(self.history_list))
        for row, data in enumerate(self.history_list):
            self.table_widget.setItem(row, 0, QTableWidgetItem(data["numero_trabajo"]))
            self.table_widget.setItem(row, 1, QTableWidgetItem(data["anio_trabajo"]))
            self.table_widget.setItem(row, 2, QTableWidgetItem(data["estado_anterior"]))
            self.table_widget.setItem(row, 3, QTableWidgetItem(data["estado_nuevo"]))
            self.table_widget.setItem(row, 4, QTableWidgetItem(data["datetime"]))

    def update_save_button_state(self):
        if self.table_widget.rowCount() > 0:
            self.save_button.setEnabled(True)
        else:
            self.save_button.setEnabled(False)

    def get_default_filename(self):
        return datetime.now().strftime('%d-%m-%Y_%H-%M')+'_historial'

    def save_to_excel(self):
        default_filename = f"{self.get_default_filename()}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(self, "Guardar como", default_filename, "Excel Files (*.xlsx);;All Files (*)")

        if not file_path:
            return

        if not file_path.endswith('.xlsx'):
            file_path += ".xlsx"

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Historial"
            headers = ["Número de Trabajo", "Año", "Estado Anterior", "Estado Nuevo", "Fecha y Hora"]
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col).value = header

            for row in range(self.table_widget.rowCount()):
                for col in range(self.table_widget.columnCount()):
                    sheet.cell(row=row + 2, column=col + 1).value = self.table_widget.item(row, col).text()

            workbook.save(file_path)
            QMessageBox.information(self, "Éxito", f"Historial guardado exitosamente en {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo guardar el archivo: {str(e)}")

    def closeEvent(self, event):
        super().closeEvent(event)
